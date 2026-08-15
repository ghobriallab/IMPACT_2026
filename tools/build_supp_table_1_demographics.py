#!/usr/bin/env python3
# ============================================================================
# Purpose:      Build Supplementary Table 1, the demographic and clinical
#               characteristics of the ELISA serological cohort (n=731) by
#               disease group, with a test of whether each characteristic
#               differs between each disease group and healthy donors.
#
#               Reviewers asked to see not only the distribution of each
#               characteristic but whether it differs by disease stage. HD is
#               the reference group for every cross-sectional comparison in the
#               paper, so each disease group is tested against HD rather than
#               all-versus-all.
#
#               Every categorical variable lists all of its levels, so the
#               counts in each column sum to that group's n. Age additionally
#               reports how many participants have a date of birth on record.
#
#               Tests, also named per variable in the "Statistical test" column:
#                 continuous   Two-sided Wilcoxon rank-sum test, equivalently
#                              the Mann-Whitney U test (scipy.stats.mannwhitneyu,
#                              alternative="two-sided"). This is the UNPAIRED
#                              test: the groups are different individuals. It is
#                              NOT the Wilcoxon signed-rank test, which is the
#                              paired version used elsewhere in this study for
#                              pre versus post vaccination comparisons within
#                              the same participants. With ties present at these
#                              sample sizes scipy uses the normal approximation
#                              with continuity and tie corrections.
#                 categorical  Fisher's exact test on the 2x2 table, two-sided,
#                              conditional on the margins. Race is tested as
#                              White versus all other categories combined.
#
#               MULTIPLE TESTING: each characteristic is tested three times,
#               once per disease group against HD, so the reported values are
#               Benjamini-Hochberg q-values computed within each characteristic
#               across those three comparisons. Cells reaching q<0.1, the
#               significance threshold used throughout this study, are printed
#               in red, matching the caption.
#
#               All 731 participants received a second mRNA dose, so age at the
#               second dose is defined cohort-wide; the participants whose
#               sample was drawn between the first and second dose still have a
#               recorded second-dose date. Age is missing only where date of
#               birth was unavailable, reported as "Unknown".
#
#               IgM-MGUS participants are reported within MGUS here to match
#               the grouping used in Figure 1; they are kept separate in
#               Figure 5C, which follows the published cohort definition.
#
# Inputs:       data/elisa/elisa_cohort_demographics.csv
#                 one de-identified row per participant. Age is age at the
#                 second vaccine dose, floored at 18 and capped at 90 per HIPAA
#                 safe harbour, so the maximum printed age is 90.
# Outputs:      tables/Supplementary_Table_1_Demographics.xlsx
# Dependencies: python3 + pandas, numpy, scipy, statsmodels, openpyxl.
# ============================================================================
from pathlib import Path
import numpy as np
import pandas as pd
from scipy import stats
from statsmodels.stats.multitest import multipletests
from openpyxl.styles import Font

REPO = Path(__file__).resolve().parent.parent
SRC = REPO / "data" / "elisa" / "elisa_cohort_demographics.csv"
OUT = REPO / "tables" / "Supplementary_Table_1_Demographics.xlsx"
OUT.parent.mkdir(parents=True, exist_ok=True)

GROUPS = ["HD", "MGUS", "SMM", "MM"]
CAPTION = ("Supplementary Table 1. Demographic and clinical characteristics of the ELISA "
           "serological cohort (n=731) by disease group. p-values were Benjamini-Hochberg "
           "corrected across the three comparisons within each characteristic. Significant "
           "q-values are colored in red.")
Q_SIGNIF = 0.1          # study-wide significance threshold
SCI_BELOW = 1e-4        # below this, cells are formatted in scientific notation
RED = Font(color="FFFF0000")
BOLD = Font(bold=True)
COL_WIDTHS = {"A": 30.83203125, "B": 10.5, "C": 12.83203125, "D": 11.83203125,
              "E": 9.83203125, "F": 37.0, "G": 22.6640625, "H": 21.6640625, "I": 20.5}

d = pd.read_csv(SRC)
d["Group"] = d["Diagnosis"].replace({"Healthy": "HD", "IgM-MGUS": "MGUS"})
d = d[d["Group"].isin(GROUPS)]
g = {k: d[d["Group"] == k] for k in GROUPS}
n = {k: len(g[k]) for k in GROUPS}

rows = []


def corrected(p_raw):
    """BH across the three disease-group comparisons of one characteristic."""
    q = multipletests(p_raw, method="fdr_bh")[1]
    return [float(f"{v:.1e}") if v < 0.001 else float(f"{v:.3f}") for v in q]


def header(label, test, p_raw):
    rows.append([label, "", "", "", "", test] + corrected(p_raw))


def level(label, cells):
    rows.append([f"    {label}"] + cells + ["", "", "", ""])


def continuous(label, col, missing_label):
    v = {k: pd.to_numeric(g[k][col], errors="coerce").dropna() for k in GROUPS}
    p_raw = [stats.mannwhitneyu(v["HD"], v[k], alternative="two-sided").pvalue for k in GROUPS[1:]]
    header(label, "Two-sided Wilcoxon rank-sum test vs HD", p_raw)
    level("Median (IQR)", [f"{v[k].median():.0f} ({v[k].quantile(.25):.0f}-{v[k].quantile(.75):.0f})"
                           for k in GROUPS])
    level("Data available, n", [f"{len(v[k])} ({100 * len(v[k]) / n[k]:.0f}%)" for k in GROUPS])
    level(missing_label, [f"{n[k] - len(v[k])} ({100 * (n[k] - len(v[k])) / n[k]:.0f}%)"
                          for k in GROUPS])


def categorical(label, series_fn, levels, test_note=None, binary_ref=None):
    """levels maps the value in the data to the label printed in the table."""
    s = {k: series_fn(g[k]) for k in GROUPS}
    counts = {lv: {k: int((s[k] == lv).sum()) for k in GROUPS} for lv in levels}
    ref = binary_ref if binary_ref is not None else next(iter(levels))
    hit = {k: int((s[k] == ref).sum()) for k in GROUPS}
    p_raw = [stats.fisher_exact([[hit["HD"], n["HD"] - hit["HD"]],
                                 [hit[k], n[k] - hit[k]]]).pvalue for k in GROUPS[1:]]
    header(label, test_note or "Fisher's exact test vs HD", p_raw)
    for lv, shown in levels.items():
        level(shown, [f"{counts[lv][k]} ({100 * counts[lv][k] / n[k]:.0f}%)" for k in GROUPS])


def race_level(x):
    r = x["Race"].astype(str)
    out = pd.Series("More than one race", index=r.index)
    out[r == "White"] = "White"
    out[r == "Black"] = "Black"
    out[r == "Asian"] = "Asian"
    out[r == "Others"] = "Other"
    out[r == "Not_Answered"] = "Unknown"
    return out


continuous("Age at second vaccine dose, years", "Age_at_second_dose", "Unknown, n")
categorical("Sex, n (%)", lambda x: x["Sex"], {"Female": "Female", "Male": "Male"})
categorical("Race, n (%)", race_level,
            {"White": "White", "Black": "Black", "Asian": "Asian",
             "More than one race": "More than one race", "Other": "Other",
             "Unknown": "Unknown, n"},
            test_note="Fisher's exact test vs HD (White vs all other)", binary_ref="White")
categorical("Vaccine product, n (%)", lambda x: x["Vaccine_type"],
            {"Pfizer": "Pfizer", "Moderna": "Moderna"})

cols = (["Characteristic"] + [f"{k} (n={n[k]})" for k in GROUPS] +
        ["Statistical test", "MGUS vs HD (q-value)",
         "SMM vs HD (q-value)", "MM vs HD (q-value)"])
table = pd.DataFrame(rows, columns=cols)

sheet = OUT.stem[:31]
with pd.ExcelWriter(OUT, engine="openpyxl") as xl:
    table.to_excel(xl, sheet_name=sheet, index=False, startrow=2)
    ws = xl.sheets[sheet]
    ws.cell(row=1, column=1, value=CAPTION)
    for letter, width in COL_WIDTHS.items():
        ws.column_dimensions[letter].width = width
    for col in range(1, len(cols) + 1):            # header row
        ws.cell(row=3, column=col).font = BOLD
    for i in range(len(table)):                       # q columns are G, H, I
        for col in (7, 8, 9):
            cell = ws.cell(row=4 + i, column=col)
            if not isinstance(cell.value, (int, float)):
                continue
            if cell.value < SCI_BELOW:
                cell.number_format = "0.00E+00"
            if cell.value < Q_SIGNIF:
                cell.font = RED

print(f"wrote {OUT}")
print(table.fillna("").to_string(index=False))
